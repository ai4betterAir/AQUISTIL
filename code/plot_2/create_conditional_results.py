#!/usr/bin/env python3
"""Create pollutant-wise conditional-formatting Excel summaries.

Outputs are written next to regional_pooled_metrics.csv:
- Persite_conditional_results.xlsx from Scope == "Site"
- perregion_conditional_results.xlsx from Scope == "Region_Micro"
- perallregion_conditional_results.xlsx from Scope == "Region_Macro"
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
AQUISTIL_ROOT = SCRIPT_DIR.parent
DEFAULT_METRICS_CSV = (
    AQUISTIL_ROOT
    / "Outputs/Imputation_Result/Metrics/regional_pooled_metrics.csv"
)

OUTPUTS = {
    "Site": "Persite_conditional_results.xlsx",
    "Region_Micro": "perregion_conditional_results.xlsx",
    "Region_Macro": "perallregion_conditional_results.xlsx",
}

COLUMNS = [
    "Region",
    "Site",
    "Target",
    "Model",
    "Regime",
    "Missingness_Percent",
    "RMSE",
    "R",
]

GROUP_COLUMNS = ["Region", "Site", "Target", "Regime", "Missingness_Percent"]
BLANK_REPEAT_COLUMNS = ["Region", "Site", "Regime", "Missingness_Percent"]


def safe_sheet_name(value: object) -> str:
    return re.sub(r'[:\\/?*\[\]]', "_", str(value))[:31]


def make_display_rows(data):
    sheets = {}
    for pollutant, pollutant_rows in data.groupby("Target", sort=True):
        rows = []
        for _, block in pollutant_rows.groupby(GROUP_COLUMNS, sort=False, dropna=False):
            block = block.copy()
            if len(block) > 1:
                block.loc[block.index[1:], BLANK_REPEAT_COLUMNS] = ""
            rows.extend(block.to_dict("records"))
            rows.append({column: "" for column in COLUMNS})
        sheets[safe_sheet_name(pollutant)] = pd.DataFrame(rows, columns=COLUMNS)
    return sheets


def apply_workbook_formatting(path: Path) -> None:
    workbook = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_black = Side(style="thin", color="000000")
    cell_border = Border(
        left=thin_black, right=thin_black, top=thin_black, bottom=thin_black
    )

    rmse_rule = ColorScaleRule(
        start_type="min",
        start_color="63BE7B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="F8696B",
    )
    r_rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    )

    widths = {
        "Region": 24,
        "Site": 22,
        "Target": 10,
        "Model": 18,
        "Regime": 18,
        "Missingness_Percent": 22,
        "RMSE": 12,
        "R": 12,
    }

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        headers = [cell.value for cell in sheet[1]]
        column_index = {name: index + 1 for index, name in enumerate(headers)}

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        for name, width in widths.items():
            sheet.column_dimensions[get_column_letter(column_index[name])].width = width

        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, max_col=sheet.max_column
        ):
            is_blank = all(cell.value in (None, "") for cell in row)
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if not is_blank:
                    cell.border = cell_border
            if not is_blank:
                sheet.cell(row=row[0].row, column=column_index["RMSE"]).number_format = (
                    "0.0000"
                )
                sheet.cell(row=row[0].row, column=column_index["R"]).number_format = (
                    "0.0000"
                )
                sheet.cell(
                    row=row[0].row, column=column_index["Missingness_Percent"]
                ).number_format = "0.00"

        starts = [
            row
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, column_index["Target"]).value not in (None, "")
        ]
        starts.append(sheet.max_row + 2)
        rmse_letter = get_column_letter(column_index["RMSE"])
        r_letter = get_column_letter(column_index["R"])
        for start, next_start in zip(starts, starts[1:]):
            end = next_start - 2
            if end > start:
                sheet.conditional_formatting.add(
                    f"{rmse_letter}{start}:{rmse_letter}{end}", rmse_rule
                )
                sheet.conditional_formatting.add(
                    f"{r_letter}{start}:{r_letter}{end}", r_rule
                )

    workbook.save(path)


def write_conditional_workbook(data: pd.DataFrame, output_path: Path) -> None:
    sheets = make_display_rows(data)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_data in sheets.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
    apply_workbook_formatting(output_path)


def main() -> None:
    metrics = pd.read_csv(DEFAULT_METRICS_CSV)
    metrics = metrics.loc[:, COLUMNS + ["Scope"]].copy()
    metrics["Missingness_Percent"] = pd.to_numeric(
        metrics["Missingness_Percent"], errors="coerce"
    )

    output_dir = DEFAULT_METRICS_CSV.parent
    for scope, filename in OUTPUTS.items():
        subset = metrics.loc[metrics["Scope"].eq(scope), COLUMNS].copy()
        if subset.empty:
            print(f"Skipped {filename}: no rows for Scope={scope}")
            continue
        if scope != "Site":
            subset["Site"] = ""
        subset = subset.sort_values(
            ["Region", "Site", "Regime", "Missingness_Percent", "Target", "Model"],
            kind="stable",
        )
        output_path = output_dir / filename
        write_conditional_workbook(subset, output_path)
        print(f"{output_path} ({len(subset):,} data rows)")


if __name__ == "__main__":
    main()
